# src/business_decision/procurement_report.py
# -*- coding: utf-8 -*-
"""
采购建议业务报表（运营版）

功能：
- 读取 procurement_suggestion.csv
- 生成 Excel 报表（按品类汇总 / 按门店汇总 / 明细）
- Excel 自动美化（列宽、千分位、异常高亮）
- 生成 PDF 周报风格报告
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INPUT_PATH = Path("data/output/procurement_suggestion.csv")
EXCEL_PATH = Path("data/output/procurement_business_report.xlsx")
PDF_PATH = Path("reports/procurement_business_report.pdf")


# ---------- 公共 ----------

def setup_chinese_font():
    plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "Arial Unicode MS"]
    plt.rcParams["axes.unicode_minus"] = False


def load_data() -> pd.DataFrame:
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"找不到采购建议文件：{INPUT_PATH}")
    df = pd.read_csv(INPUT_PATH)

    # 数值列
    num_cols = [
        "forecast_demand_1d",
        "forecast_demand_3d",
        "avg_daily_demand",
        "stock_on_hand",
        "stock_in_transit",
        "required_stock_level",
        "recommended_purchase_qty",
        "purchase_qty_rounded",
        "y_price_1d_pred",
    ]
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    for col in ["run_date", "target_date", "ts"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.date

    return df


# ---------- 汇总：按品类 ----------

def add_class_level_flags(df: pd.DataFrame) -> pd.DataFrame:
    """按品类汇总（增强版 BI 风格）。"""
    df = df.copy()
    if "classify_name" not in df.columns:
        df["classify_name"] = "未分类"

    # 二级品类：这里简单用 variety 占位，后续你可以替换成映射
    if "variety" in df.columns:
        df["sub_category"] = df["variety"]
    else:
        df["sub_category"] = "未知"

    sku_count = (
        df.groupby("classify_name")["product_id"]
        .nunique()
        .rename("sku_count")
    )

    agg = df.groupby("classify_name").agg(
        demand_1d=("forecast_demand_1d", "sum"),
        demand_3d=("forecast_demand_3d", "sum"),
        stock_on_hand=("stock_on_hand", "sum"),
        stock_in_transit=("stock_in_transit", "sum"),
        purchase_qty=("purchase_qty_rounded", "sum"),
    )
    agg = agg.join(sku_count)

    agg["avg_daily_demand"] = agg["demand_3d"] / 3.0

    peak = (
        df.groupby("classify_name")["forecast_demand_3d"]
        .max()
        .rename("peak_demand")
    )
    agg = agg.join(peak)

    cv = (
        df.groupby("classify_name")["forecast_demand_3d"]
        .std()
        .rename("demand_cv")
    )
    agg = agg.join(cv.fillna(0.0))

    agg["available_stock"] = agg["stock_on_hand"] + agg["stock_in_transit"]
    agg["coverage_days"] = np.where(
        agg["avg_daily_demand"] > 0,
        agg["available_stock"] / agg["avg_daily_demand"],
        np.inf,
    )

    total_purchase = agg["purchase_qty"].sum()
    agg["purchase_share"] = agg["purchase_qty"] / (total_purchase + 1e-9)
    agg["avg_purchase_per_sku"] = agg["purchase_qty"] / (agg["sku_count"] + 1e-9)

    warnings = []
    for _, row in agg.iterrows():
        w = []
        if row["demand_3d"] == 0 and row["purchase_qty"] > 0:
            w.append("无需求却仍采购")
        if row["demand_3d"] > 0 and row["purchase_qty"] == 0:
            w.append("有需求却未采购")
        if row["coverage_days"] > 7 and np.isfinite(row["coverage_days"]):
            w.append("库存覆盖>7天，疑似压货")
        if 0 < row["coverage_days"] < 1 and row["avg_daily_demand"] > 0:
            w.append("库存覆盖<1天，疑似断货")
        if row["demand_cv"] > 1.2:
            w.append("需求波动大（CV>1.2）")
        if (
            row["avg_daily_demand"] > 0
            and row["peak_demand"] > row["avg_daily_demand"] * 3
        ):
            w.append("峰值需求超过平均3倍，波动过大")

        warnings.append("；".join(w) if w else "正常")

    agg["warning"] = warnings

    agg = agg.sort_values(
        ["purchase_qty", "demand_cv", "coverage_days"],
        ascending=[False, False, True],
    ).reset_index()

    agg_out = agg.rename(
        columns={
            "classify_name": "品类",
            "demand_1d": "未来1天需求",
            "demand_3d": "未来3天需求",
            "avg_daily_demand": "平均日需求",
            "peak_demand": "峰值需求",
            "demand_cv": "需求波动系数CV",
            "stock_on_hand": "当前库存",
            "stock_in_transit": "在途库存",
            "available_stock": "可用库存",
            "coverage_days": "库存覆盖天数",
            "purchase_qty": "建议采购量",
            "purchase_share": "采购占比",
            "sku_count": "SKU数量",
            "avg_purchase_per_sku": "单SKU平均采购量",
            "warning": "异常提醒",
        }
    )

    return agg_out


# ---------- 汇总：按门店 ----------

def add_shop_level_flags(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "shop_name" not in df.columns:
        df["shop_name"] = "未知门店"

    agg = df.groupby("shop_name").agg(
        demand_1d=("forecast_demand_1d", "sum"),
        demand_3d=("forecast_demand_3d", "sum"),
        stock_on_hand=("stock_on_hand", "sum"),
        stock_in_transit=("stock_in_transit", "sum"),
        purchase_qty=("purchase_qty_rounded", "sum"),
    )

    agg["avg_daily_demand"] = agg["demand_3d"] / 3.0
    agg["available_stock"] = agg["stock_on_hand"] + agg["stock_in_transit"]
    agg["coverage_days"] = np.where(
        agg["avg_daily_demand"] > 0,
        agg["available_stock"] / agg["avg_daily_demand"],
        np.inf,
    )

    total_purchase = agg["purchase_qty"].sum()
    agg["purchase_share"] = agg["purchase_qty"] / (total_purchase + 1e-9)

    warnings = []
    for _, row in agg.iterrows():
        w = []
        if row["demand_3d"] == 0 and row["purchase_qty"] > 0:
            w.append("无需求却仍采购")
        if row["demand_3d"] > 0 and row["purchase_qty"] == 0:
            w.append("有需求却未采购")
        if row["coverage_days"] > 7 and np.isfinite(row["coverage_days"]):
            w.append("库存覆盖>7天，疑似压货")
        if 0 < row["coverage_days"] < 1 and row["avg_daily_demand"] > 0:
            w.append("库存覆盖<1天，疑似断货")

        warnings.append("；".join(w) if w else "正常")

    agg["warning"] = warnings

    agg = agg.sort_values("purchase_qty", ascending=False).reset_index()

    agg_out = agg.rename(
        columns={
            "shop_name": "门店",
            "demand_1d": "未来1天需求",
            "demand_3d": "未来3天需求",
            "avg_daily_demand": "平均日需求",
            "stock_on_hand": "当前库存",
            "stock_in_transit": "在途库存",
            "available_stock": "可用库存",
            "coverage_days": "库存覆盖天数",
            "purchase_qty": "建议采购量",
            "purchase_share": "采购占比",
            "warning": "异常提醒",
        }
    )
    return agg_out


# ---------- 明细表 ----------

def make_detail_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    业务友好版采购明细（以 variety + grade 为主导）
    并新增 product_id 字段。

    输出结构：

    | 品种 | 等级 | 产品ID | 门店 | 规格 | 品类 | 未来3天需求 | 建议采购量 | 预测单价 | 价格信号 | 备注 |
    """
    col_map = []

    # ---- 主导信息 ----
    col_map.append(("variety", "品种", True))
    col_map.append(("grade", "等级", True))
    col_map.append(("product_id", "产品ID", False))   # 新增字段！

    # ---- 门店维度 ----
    col_map.append(("shop_name", "门店", True))

    # ---- 产品描述 ----
    col_map.append(("spec", "规格", False))
    col_map.append(("classify_name", "品类", False))

    # ---- 数值类字段 ----
    col_map.append(("forecast_demand_3d", "未来3天需求", False))
    col_map.append(("purchase_qty_rounded", "建议采购量", False))
    col_map.append(("y_price_1d_pred", "预测单价", False))

    # ---- 信号 & 备注 ----
    col_map.append(("price_signal", "价格信号", False))
    col_map.append(("decision_reason", "备注", False))

    # ---- 构建最终列 ----
    exist_src_cols = []
    exist_dst_cols = []
    sort_keys = []

    for src, dst, is_sort_key in col_map:
        if src in df.columns:
            exist_src_cols.append(src)
            exist_dst_cols.append(dst)
            if is_sort_key:
                sort_keys.append(dst)

    detail = df[exist_src_cols].copy()
    detail.rename(columns=dict(zip(exist_src_cols, exist_dst_cols)), inplace=True)

    # ---- 排序逻辑：品种 -> 等级 -> 门店 -> 建议采购量(降序）----
    sort_cols = []
    for key in ["品种", "等级", "门店"]:
        if key in detail.columns:
            sort_cols.append(key)
    if "建议采购量" in detail.columns:
        sort_cols.append("建议采购量")

    if sort_cols:
        ascending = [True] * (len(sort_cols) - 1) + [False]
        detail.sort_values(sort_cols, ascending=ascending, inplace=True)

    return detail



# ---------- Excel 美化 ----------

def format_sheet(ws, has_warning_col: bool = True):
    """简单美化：标题加粗、自动列宽、数字格式、异常高亮。"""
    # 冻结首行
    ws.freeze_panes = "A2"

    # 边框 & 样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 标题行样式
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 自动列宽 + 内容样式
    col_max = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        for idx, value in enumerate(row, start=1):
            length = len(str(value)) if value is not None else 0
            col_max[idx] = max(col_max.get(idx, 0), length)

    for idx, width in col_max.items():
        ws.column_dimensions[get_column_letter(idx)].width = min(width + 2, 40)

    # 如果有“异常提醒”列，标红非“正常”的行
    warning_col_idx = None
    if has_warning_col:
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value == "异常提醒":
                warning_col_idx = idx
                break

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
        if warning_col_idx is not None:
            warning_val = row[warning_col_idx - 1].value
            if isinstance(warning_val, str) and warning_val != "正常":
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="F8CBAD")


def export_excel(df: pd.DataFrame):
    df_class = add_class_level_flags(df)
    df_shop = add_shop_level_flags(df)
    df_detail = make_detail_table(df)

    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl") as writer:
        df_class.to_excel(writer, sheet_name="按品类汇总", index=False)
        df_shop.to_excel(writer, sheet_name="按门店汇总", index=False)
        df_detail.to_excel(writer, sheet_name="采购明细", index=False)

        wb = writer.book
        for name in ["按品类汇总", "按门店汇总", "采购明细"]:
            ws = writer.sheets[name]
            format_sheet(ws, has_warning_col=(name != "采购明细"))

    print(f"✅ 已生成 Excel 报告：{EXCEL_PATH}")


# ---------- PDF 运营版报告 ----------

def export_pdf(df: pd.DataFrame):
    setup_chinese_font()

    df_class = add_class_level_flags(df)
    df_shop = add_shop_level_flags(df)

    total_purchase = df["purchase_qty_rounded"].sum()
    total_sku = df.shape[0]
    run_date = df["run_date"].max() if "run_date" in df.columns else None

    top_class = df_class.sort_values("建议采购量", ascending=False).head(10)
    top_shop = df_shop.sort_values("建议采购量", ascending=False).head(10)

    risky_class = df_class[df_class["异常提醒"] != "正常"].copy()
    risky_shop = df_shop[df_shop["异常提醒"] != "正常"].copy()

    PDF_PATH.parent.mkdir(parents=True, exist_ok=True)

    with PdfPages(PDF_PATH) as pdf:
        # 封面
        plt.figure(figsize=(8.27, 11.69))  # A4
        plt.axis("off")
        title = "鲜花采购建议日报"
        sub_title = f"运行日期：{run_date}" if run_date else "运行日期：未知"
        kpi1 = f"总 SKU 数量：{total_sku:,}"
        kpi2 = f"总建议采购量：{total_purchase:,.0f}"

        plt.text(0.5, 0.8, title, fontsize=24, ha="center")
        plt.text(0.5, 0.7, sub_title, fontsize=14, ha="center")
        plt.text(0.5, 0.6, kpi1, fontsize=12, ha="center")
        plt.text(0.5, 0.55, kpi2, fontsize=12, ha="center")
        pdf.savefig()
        plt.close()

        # 品类 Top10
        plt.figure(figsize=(8.27, 11.69))
        plt.subplot(2, 1, 1)
        plt.title("按品类建议采购量 Top10")
        plt.bar(top_class["品类"], top_class["建议采购量"])
        plt.xticks(rotation=45, ha="right")
        plt.ylabel("建议采购量")

        plt.subplot(2, 1, 2)
        plt.title("按品类采购占比 Top10")
        plt.bar(top_class["品类"], top_class["采购占比"])
        plt.xticks(rotation=45, ha="right")
        plt.ylabel("占比")
        plt.tight_layout()
        pdf.savefig()
        plt.close()

        # 门店 Top10
        plt.figure(figsize=(8.27, 11.69))
        plt.subplot(2, 1, 1)
        plt.title("按门店建议采购量 Top10")
        plt.bar(top_shop["门店"], top_shop["建议采购量"])
        plt.xticks(rotation=45, ha="right")
        plt.ylabel("建议采购量")

        plt.subplot(2, 1, 2)
        plt.title("按门店库存覆盖天数 Top10")
        plt.bar(top_shop["门店"], top_shop["库存覆盖天数"])
        plt.xticks(rotation=45, ha="right")
        plt.ylabel("天数")
        plt.tight_layout()
        pdf.savefig()
        plt.close()

        # 风险页：品类 & 门店
        plt.figure(figsize=(8.27, 11.69))
        plt.axis("off")
        plt.title("风险提示（品类维度）", fontsize=14, loc="left")
        y = 0.9
        if risky_class.empty:
            plt.text(0.05, y, "当前未发现品类级异常。", fontsize=11)
        else:
            for _, row in risky_class.iterrows():
                line = f"- {row['品类']}: {row['异常提醒']}"
                plt.text(0.05, y, line, fontsize=10)
                y -= 0.05
                if y < 0.5:
                    break

        plt.text(0.05, 0.45, "风险提示（门店维度）", fontsize=14)
        y = 0.4
        if risky_shop.empty:
            plt.text(0.05, y, "当前未发现门店级异常。", fontsize=11)
        else:
            for _, row in risky_shop.iterrows():
                line = f"- {row['门店']}: {row['异常提醒']}"
                plt.text(0.05, y, line, fontsize=10)
                y -= 0.05
                if y < 0.1:
                    break

        pdf.savefig()
        plt.close()

    print(f"✅ 已生成 PDF 报告：{PDF_PATH}")


def main():
    df = load_data()
    export_excel(df)
    export_pdf(df)


if __name__ == "__main__":
    main()
